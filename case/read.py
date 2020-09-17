# 读取Excel测试数据
import openpyxl

workbook = openpyxl.load_workbook("../data/buy_safe.xlsx")
print(workbook)
sheetnames = workbook.sheetnames
print(sheetnames)
worksheet = workbook[sheetnames[0]]
print(worksheet)

for row in list(worksheet.rows)[1:]:
    type = row[0].value
    year = row[1].value
    expect = row[2].value
    print(type, year, expect)